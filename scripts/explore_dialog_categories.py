"""
대화체 파일의 분류 체계(대분류/소분류/상황) 전체 고유값 추출

실행: python scripts/explore_dialog_categories.py
출력: scripts/dialog_categories.txt
"""

import sys
from pathlib import Path
from collections import Counter

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "docs" / "한국어-영어 번역(병렬) 말뭉치"
OUTPUT = Path(__file__).parent / "dialog_categories.txt"

lines = []

def log(msg=""):
    lines.append(msg)
    print(msg)


def main():
    filepath = DATA_DIR / "2_대화체.xlsx"
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    # 계층별 수집
    major_counter = Counter()          # 대분류
    minor_counter = Counter()          # 대분류 > 소분류
    situation_counter = Counter()      # 대분류 > 소분류 > 상황
    set_count = set()

    for row in ws.iter_rows(min_row=2, values_only=True):
        major = str(row[0]).strip() if row[0] else ""
        minor = str(row[1]).strip() if row[1] else ""
        situation = str(row[2]).strip() if row[2] else ""
        set_nr = row[3]

        major_counter[major] += 1
        minor_counter[f"{major} > {minor}"] += 1
        situation_counter[f"{major} > {minor} > {situation}"] += 1
        if set_nr:
            set_count.add(set_nr)

    wb.close()

    log("=" * 70)
    log("대화체 파일 분류 체계")
    log("=" * 70)

    log(f"\n총 대화 세트 수: {len(set_count):,}개")
    log(f"총 발화 수: {sum(major_counter.values()):,}문장")

    log(f"\n{'─'*70}")
    log(f"[대분류] ({len(major_counter)}종)")
    log(f"{'─'*70}")
    for cat, count in major_counter.most_common():
        log(f"  {cat}: {count:,}문장")

    log(f"\n{'─'*70}")
    log(f"[대분류 > 소분류] ({len(minor_counter)}종)")
    log(f"{'─'*70}")
    current_major = ""
    for key, count in sorted(minor_counter.items()):
        major = key.split(" > ")[0]
        if major != current_major:
            log(f"\n  [{major}]")
            current_major = major
        minor = key.split(" > ")[1]
        log(f"    {minor}: {count:,}문장")

    log(f"\n{'─'*70}")
    log(f"[대분류 > 소분류 > 상황] ({len(situation_counter)}종)")
    log(f"{'─'*70}")
    current_minor = ""
    for key, count in sorted(situation_counter.items()):
        parts = key.split(" > ")
        minor_key = f"{parts[0]} > {parts[1]}"
        if minor_key != current_minor:
            log(f"\n  [{parts[0]} > {parts[1]}]")
            current_minor = minor_key
        log(f"    {parts[2]}: {count:,}문장")

    OUTPUT.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n저장 완료: {OUTPUT}")


if __name__ == "__main__":
    main()
