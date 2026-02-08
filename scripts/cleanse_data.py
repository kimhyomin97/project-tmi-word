"""
AI Hub 한영 병렬 말뭉치 정제 스크립트 (v2)
- 5단계 난이도: STARTER / BEGINNER / INTERMEDIATE / ADVANCED / CHALLENGE
- 구어체(1), 구어체(2), 대화체 파일에서 학습 앱에 적합한 문장을 필터링
- 결과를 CSV로 출력

실행: python scripts/cleanse_data.py
출력: scripts/cleansed_sentences.csv, scripts/cleanse_report.txt
"""

import csv
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "docs" / "한국어-영어 번역(병렬) 말뭉치"
OUTPUT_CSV = Path(__file__).parent / "cleansed_sentences.csv"
REPORT_FILE = Path(__file__).parent / "cleanse_report.txt"

# ──────────────────────────────────────────────
# 필터링 기준 (v2: 길이 범위 확장)
# ──────────────────────────────────────────────
EN_WORD_MIN = 1       # 영어 최소 단어 수 (1단어도 일단 통과 → STARTER로 분류)
EN_WORD_MAX = 35      # 영어 최대 단어 수 (CHALLENGE 구간까지 포함)
KO_CHAR_MIN = 2       # 한글 최소 글자 수 (완화)
KO_CHAR_MAX = 150     # 한글 최대 글자 수 (완화)
SPECIAL_CHAR_RATIO = 0.25  # 특수문자/숫자 비율 상한

# 5단계 난이도 기준 (영어 단어 수 기반)
# STARTER:       1~3단어   짧은 표현, 관용구
# BEGINNER:      4~8단어   기본 문장
# INTERMEDIATE:  9~15단어  일반 회화
# ADVANCED:      16~25단어 복문, 관계절
# CHALLENGE:     26~35단어 다중 종속절, 긴 복합문
# ──────────────────────────────────────────────

report_lines = []


def log(msg=""):
    report_lines.append(msg)
    print(msg)


def count_en_words(text: str) -> int:
    return len(text.split())


def special_char_ratio(text: str) -> float:
    """특수문자 + 숫자 비율"""
    if not text:
        return 0
    special = sum(1 for c in text if not c.isalpha() and not c.isspace())
    return special / len(text)


def has_mixed_lang(en_text: str, ko_text: str) -> bool:
    """영어 문장에 한글이 섞여있거나, 한글 문장이 영어로만 되어있는지"""
    ko_in_en = re.findall(r'[가-힣]+', en_text)
    if sum(len(w) for w in ko_in_en) > 4:
        return True
    if not re.search(r'[가-힣]', ko_text):
        return True
    return False


def classify_difficulty(en_text: str) -> str:
    word_count = count_en_words(en_text)
    if word_count <= 3:
        return "STARTER"
    elif word_count <= 8:
        return "BEGINNER"
    elif word_count <= 15:
        return "INTERMEDIATE"
    elif word_count <= 25:
        return "ADVANCED"
    else:
        return "CHALLENGE"


def filter_sentence(en_text: str, ko_text: str) -> tuple[bool, str]:
    """
    필터 통과 여부와 탈락 사유를 반환.
    Returns: (통과 여부, 사유)
    """
    if not en_text or not ko_text:
        return False, "empty"

    en_text = str(en_text).strip()
    ko_text = str(ko_text).strip()

    if not en_text or not ko_text:
        return False, "empty_after_strip"

    en_words = count_en_words(en_text)
    if en_words < EN_WORD_MIN:
        return False, "en_too_short"
    if en_words > EN_WORD_MAX:
        return False, "en_too_long"

    ko_len = len(ko_text)
    if ko_len < KO_CHAR_MIN:
        return False, "ko_too_short"
    if ko_len > KO_CHAR_MAX:
        return False, "ko_too_long"

    if special_char_ratio(en_text) > SPECIAL_CHAR_RATIO:
        return False, "en_special_chars"
    if special_char_ratio(ko_text) > SPECIAL_CHAR_RATIO:
        return False, "ko_special_chars"

    if has_mixed_lang(en_text, ko_text):
        return False, "mixed_lang"

    return True, "ok"


def process_spoken_file(filepath: Path, source_label: str) -> list[dict]:
    """구어체 파일 처리 (컬럼: SID, 원문, 번역문)"""
    log(f"\n처리 중: {filepath.name}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    results = []
    stats = {"total": 0, "passed": 0, "reasons": {}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["total"] += 1
        ko_text = str(row[1]).strip() if row[1] else ""
        en_text = str(row[2]).strip() if row[2] else ""

        passed, reason = filter_sentence(en_text, ko_text)
        if passed:
            stats["passed"] += 1
            results.append({
                "english_text": en_text,
                "korean_ref": ko_text,
                "difficulty": classify_difficulty(en_text),
                "category": "DAILY",
                "subcategory": "",
                "source": source_label,
            })
        else:
            stats["reasons"][reason] = stats["reasons"].get(reason, 0) + 1

    wb.close()

    log(f"  전체: {stats['total']:,}행 → 통과: {stats['passed']:,}행 ({stats['passed']/max(stats['total'],1)*100:.1f}%)")
    log(f"  탈락 사유:")
    for reason, count in sorted(stats["reasons"].items(), key=lambda x: -x[1]):
        log(f"    {reason}: {count:,}")

    return results


def process_dialog_file(filepath: Path) -> list[dict]:
    """대화체 파일 처리 (컬럼: 대분류, 소분류, 상황, Set Nr., 발화자, 원문, 번역문)"""
    log(f"\n처리 중: {filepath.name}")
    wb = openpyxl.load_workbook(filepath, read_only=True)
    ws = wb[wb.sheetnames[0]]

    results = []
    stats = {"total": 0, "passed": 0, "reasons": {}}

    for row in ws.iter_rows(min_row=2, values_only=True):
        stats["total"] += 1
        major_cat = str(row[0]).strip() if row[0] else ""
        minor_cat = str(row[1]).strip() if row[1] else ""
        ko_text = str(row[5]).strip() if row[5] else ""
        en_text = str(row[6]).strip() if row[6] else ""

        passed, reason = filter_sentence(en_text, ko_text)
        if passed:
            stats["passed"] += 1
            results.append({
                "english_text": en_text,
                "korean_ref": ko_text,
                "difficulty": classify_difficulty(en_text),
                "category": major_cat,
                "subcategory": minor_cat,
                "source": "AIHUB_DIALOG",
            })
        else:
            stats["reasons"][reason] = stats["reasons"].get(reason, 0) + 1

    wb.close()

    log(f"  전체: {stats['total']:,}행 → 통과: {stats['passed']:,}행 ({stats['passed']/max(stats['total'],1)*100:.1f}%)")
    log(f"  탈락 사유:")
    for reason, count in sorted(stats["reasons"].items(), key=lambda x: -x[1]):
        log(f"    {reason}: {count:,}")

    return results


def deduplicate(sentences: list[dict]) -> list[dict]:
    """영어 문장 기준 중복 제거 (대소문자 무시)"""
    seen = set()
    unique = []
    for s in sentences:
        key = s["english_text"].lower().strip()
        if key not in seen:
            seen.add(key)
            unique.append(s)
    return unique


def main():
    log("=" * 60)
    log("AI Hub 한영 병렬 말뭉치 정제 (v2 — 5단계 난이도)")
    log("=" * 60)
    log(f"영어 단어 수: {EN_WORD_MIN}~{EN_WORD_MAX}")
    log(f"한국어 글자 수: {KO_CHAR_MIN}~{KO_CHAR_MAX}")
    log(f"특수문자 비율 상한: {SPECIAL_CHAR_RATIO*100:.0f}%")
    log(f"난이도: STARTER(1-3) / BEGINNER(4-8) / INTERMEDIATE(9-15) / ADVANCED(16-25) / CHALLENGE(26-35)")

    all_sentences = []

    # 구어체 처리
    for filename, label in [
        ("1_구어체(1).xlsx", "AIHUB_SPOKEN_1"),
        ("1_구어체(2).xlsx", "AIHUB_SPOKEN_2"),
    ]:
        filepath = DATA_DIR / filename
        if filepath.exists():
            all_sentences.extend(process_spoken_file(filepath, label))
        else:
            log(f"\n파일 없음: {filename}")

    # 대화체 처리
    dialog_path = DATA_DIR / "2_대화체.xlsx"
    if dialog_path.exists():
        all_sentences.extend(process_dialog_file(dialog_path))
    else:
        log(f"\n파일 없음: 2_대화체.xlsx")

    log(f"\n{'='*60}")
    log(f"중복 제거 전: {len(all_sentences):,}문장")
    all_sentences = deduplicate(all_sentences)
    log(f"중복 제거 후: {len(all_sentences):,}문장")

    # 난이도별 통계
    diff_stats = {}
    cat_stats = {}
    source_stats = {}
    for s in all_sentences:
        diff_stats[s["difficulty"]] = diff_stats.get(s["difficulty"], 0) + 1
        cat_stats[s["category"]] = cat_stats.get(s["category"], 0) + 1
        source_stats[s["source"]] = source_stats.get(s["source"], 0) + 1

    log(f"\n난이도별 분포:")
    for d in ["STARTER", "BEGINNER", "INTERMEDIATE", "ADVANCED", "CHALLENGE"]:
        count = diff_stats.get(d, 0)
        pct = count / max(len(all_sentences), 1) * 100
        log(f"  {d:15s}: {count:>8,}  ({pct:5.1f}%)")

    log(f"\n카테고리별 분포:")
    for cat, count in sorted(cat_stats.items(), key=lambda x: -x[1]):
        pct = count / max(len(all_sentences), 1) * 100
        log(f"  {cat:15s}: {count:>8,}  ({pct:5.1f}%)")

    log(f"\n출처별 분포:")
    for src, count in sorted(source_stats.items(), key=lambda x: -x[1]):
        pct = count / max(len(all_sentences), 1) * 100
        log(f"  {src:20s}: {count:>8,}  ({pct:5.1f}%)")

    # CSV 출력
    with open(OUTPUT_CSV, "w", newline="", encoding="utf-8") as f:
        writer = csv.DictWriter(f, fieldnames=[
            "english_text", "korean_ref", "difficulty", "category", "subcategory", "source"
        ])
        writer.writeheader()
        writer.writerows(all_sentences)

    log(f"\n{'='*60}")
    log(f"CSV 저장 완료: {OUTPUT_CSV}")
    log(f"총 {len(all_sentences):,}문장")

    # 리포트 저장
    REPORT_FILE.write_text("\n".join(report_lines), encoding="utf-8")
    print(f"리포트 저장 완료: {REPORT_FILE}")


if __name__ == "__main__":
    main()
