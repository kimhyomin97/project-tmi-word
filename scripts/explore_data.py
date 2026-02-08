"""
AI Hub 한영 병렬 말뭉치 xlsx 데이터 탐색 스크립트
- 결과를 scripts/explore_result.txt 에 저장

실행: python scripts/explore_data.py
필요: pip install openpyxl
"""

import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("openpyxl이 필요합니다: pip install openpyxl")
    sys.exit(1)

DATA_DIR = Path(__file__).parent.parent / "docs" / "한국어-영어 번역(병렬) 말뭉치"
OUTPUT_FILE = Path(__file__).parent / "explore_result.txt"

PRIORITY_FILES = [
    "1_구어체(1).xlsx",
    "1_구어체(2).xlsx",
    "2_대화체.xlsx",
]

lines = []

def log(msg=""):
    lines.append(msg)
    print(msg)


def explore_xlsx(filepath: Path, sample_rows: int = 3):
    log(f"\n{'='*80}")
    log(f"파일: {filepath.name}")
    log(f"크기: {filepath.stat().st_size / (1024*1024):.1f} MB")
    log(f"{'='*80}")

    wb = openpyxl.load_workbook(filepath, read_only=True)

    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        log(f"\n  시트: {sheet_name}")
        log(f"  행 수(max_row): {ws.max_row}, 열 수: {ws.max_column}")

        headers = []
        for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
            headers = [str(h) if h else f"col_{i}" for i, h in enumerate(row)]
        log(f"  컬럼: {headers}")

        log(f"\n  --- 샘플 (첫 {sample_rows}행) ---")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=2 + sample_rows - 1, values_only=True)):
            for h, v in zip(headers, row):
                log(f"    {h}: {v}")
            log(f"    ---")

        log(f"\n  --- 마지막 2행 ---")
        last = list(ws.iter_rows(min_row=max(2, ws.max_row - 1), max_row=ws.max_row, values_only=True))
        for row in last:
            for h, v in zip(headers, row):
                log(f"    {h}: {v}")
            log(f"    ---")

    wb.close()


def main():
    if not DATA_DIR.exists():
        log(f"데이터 디렉토리를 찾을 수 없습니다: {DATA_DIR}")
        sys.exit(1)

    log(f"데이터 디렉토리: {DATA_DIR}")
    log(f"\n전체 파일 목록:")
    for f in sorted(DATA_DIR.iterdir()):
        log(f"  - {f.name} ({f.stat().st_size / (1024*1024):.1f} MB)")

    for filename in PRIORITY_FILES:
        filepath = DATA_DIR / filename
        if filepath.exists():
            explore_xlsx(filepath, sample_rows=3)
        else:
            log(f"\n  파일 없음: {filename}")

    # 결과 저장
    OUTPUT_FILE.write_text("\n".join(lines), encoding="utf-8")
    print(f"\n결과 저장 완료: {OUTPUT_FILE}")


if __name__ == "__main__":
    main()
